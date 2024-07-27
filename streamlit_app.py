import streamlit as st
import json
from datetime import datetime, timedelta
from itertools import combinations
import random
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
from icalendar import Calendar, Event, vDatetime
import pytz
from pymongo import MongoClient
import re

# Retrieve MongoDB credentials from Streamlit secrets
mongo_user = st.secrets["MONGODB"]["user"]
mongo_password = st.secrets["MONGODB"]["password"]

# MongoDB Atlas Connection
mongo_uri = f"mongodb+srv://{mongo_user}:{mongo_password}@cluster0.rkdwvgd.mongodb.net/?retryWrites=true&w=majority&appName=Cluster0"
client = MongoClient(mongo_uri)
db = client["class_schedule_db"]
collection = db["class_entries"]

# CLASS LOGGER FUNCTIONS

# Function to display a single class entry
def display_class_entry(class_entry):
    st.write(f"Nombre de la clase: {class_entry['name'].capitalize()}")
    st.write(f"Grupo/Sección: {class_entry['group']}")
    for schedule in class_entry['schedule']:
        st.write(f"Día: {schedule['day']}, Hora de inicio: {schedule['start_time']}, Hora de finalización: {schedule['end_time']}")

def save_class_to_db(id, class_data):
    collection.update_one(
        {"id": id},
        {"$push": {"classes": class_data}},
        upsert=True
    )

def get_classes_from_db(id):
    user_data = collection.find_one({"id": id})
    return user_data.get("classes", []) if user_data else []

# ----------------------------

# TIMETABLE CREATOR FUNCTIONS

# Additional functions for timetable
def parse_time(time_str):
    return datetime.strptime(time_str, '%H:%M').time()

def times_overlap(start1, end1, start2, end2):
    return max(start1, start2) < min(end1, end2)

def has_conflict(class1, class2):
    for session1 in class1['schedule']:
        for session2 in class2['schedule']:
            if session1['day'] == session2['day']:
                if times_overlap(session1['start_time'], session1['end_time'],
                                 session2['start_time'], session2['end_time']):
                    return True
    return False

def has_unique_classes(combination):
    class_names = [cls['name'] for cls in combination]
    return len(class_names) == len(set(class_names))

def has_free_days(combination, free_days):
    # Checks if the combination respects the free days
    scheduled_days = {session['day'] for cls in combination for session in cls['schedule']}
    return all(day not in scheduled_days for day in free_days)

def get_unique_time_slots(classes):
    time_slots = set()
    for cls in classes:
        for session in cls['schedule']:
            time_slots.add((session['start_time'], session['end_time']))
    return sorted(time_slots, key=lambda x: x[0])

def get_random_light_color():
    return "{:02x}{:02x}{:02x}".format(random.randint(100, 255), random.randint(100, 255), random.randint(100, 255))

def create_single_sheet_xlsx_timetables(combinations, filename, time_slots, classes):
    wb = Workbook()
    ws = wb.active
    ws.title = "Horarios"

    header_color = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    class_colors = {cls['name']: PatternFill(start_color=get_random_light_color(), end_color=get_random_light_color(), fill_type="solid") for cls in classes}

    current_row = 1
    days = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes"]

    for i, combo in enumerate(combinations):
        current_row += 1  # Move to the next row for the header

        # Set the header row and column
        for col, day in enumerate(['Hora'] + days, start=1):
            cell = ws.cell(row=current_row, column=col, value=day)
            if col != 1:  # Apply header color to days, not to the 'Time' column
                cell.fill = header_color

        # Create a dark grey border style
        dark_grey_side = Side(border_style="thin", color="404040")
        dark_grey_border = Border(left=dark_grey_side, right=dark_grey_side, top=dark_grey_side, bottom=dark_grey_side)

        # Apply the border to the header row
        for col in range(1, 7):  # Columns A-G
            ws.cell(row=current_row, column=col).border = dark_grey_border

        # Populate the timetable
        for time_slot in time_slots:
            current_row += 1
            ws.cell(row=current_row, column=1, value=f"{time_slot[0].strftime('%H:%M')} - {time_slot[1].strftime('%H:%M')}")
            ws.cell(row=current_row, column=1).fill = header_color

            for cls in combo:
                for session in cls['schedule']:
                    if (session['start_time'], session['end_time']) == time_slot:
                        day_col = days.index(session['day']) + 2
                        cell = ws.cell(row=current_row, column=day_col)
                        class_room_info = session.get('class_room', 'N/A')  # Retrieve class room or default to 'N/A'
                        cell.value = f"{cls['name']}\n(Grupo {cls['group']}\nAula: {class_room_info})"
                        cell.fill = class_colors[cls['name']]
                        cell.alignment = Alignment(wrap_text=True)

        # Apply the border to each cell in the timetable
        for row in ws.iter_rows(min_row=current_row - len(time_slots), max_row=current_row, min_col=1, max_col=6):
            for cell in row:
                cell.border = dark_grey_border

        current_row += 5 

    # Set column widths
    for i, column_width in enumerate([20] + [25] * 5, start=1):
        ws.column_dimensions[get_column_letter(i)].width = column_width

    wb.save(filename)

# ----------------------------

# CALENDAR ICS GENERATOR FUNCTIONS
    
def generate_ics_file_for_classes(selected_classes, classes, start_date_str, end_date_str, filename="horario_clases.ics"):
    cal = Calendar()
    
    start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
    end_date = datetime.strptime(end_date_str, '%Y-%m-%d')

    # Mapping Spanish day names to English day names
    day_mapping = {
        "Lunes": "Monday",
        "Martes": "Tuesday",
        "Miércoles": "Wednesday",
        "Jueves": "Thursday",
        "Viernes": "Friday",
        "Sábado": "Saturday",
        "Domingo": "Sunday"
    }

    for selected_class in selected_classes:

        cls = next(
            (item for item in classes if item["name"] == selected_class['name'] and item["group"] == selected_class['group']),
            None
        )

        if cls is None:
            continue
        
        for session in cls['schedule']:
            day, start_time_str, end_time_str = session['day'], session['start_time'], session['end_time']
            if not all([day, start_time_str, end_time_str]):
                continue

            # Translate day name to English
            day_english = day_mapping.get(day, day)

            first_occurrence_date = start_date
            # Adjust first_occurrence_date to the first occurrence of the session day
            while first_occurrence_date.strftime('%A') != day_english:
                first_occurrence_date += timedelta(days=1)
                if first_occurrence_date > end_date:
                    # If first_occurrence_date exceeds end_date, break the loop
                    break

            if first_occurrence_date > end_date:
                # If the first occurrence is beyond the semester end, skip this session
                continue

            start_datetime = datetime.combine(first_occurrence_date, datetime.strptime(start_time_str, '%H:%M').time())
            end_datetime = datetime.combine(first_occurrence_date, datetime.strptime(end_time_str, '%H:%M').time())
            if end_datetime <= start_datetime:
                # Skipping event due to invalid time range
                continue

            event = Event()
            event.add('summary', f"{cls['name']} - Grupo {cls['group']}")
            event.add('location', session.get('class_room', 'N/A'))
            event.add('dtstart', start_datetime)
            event.add('dtend', end_datetime)
            event.add('rrule', {'freq': 'weekly', 'until': end_date})
            cal.add_component(event)

    with open(filename, 'wb') as f:
        f.write(cal.to_ical())
    
    return filename


# ----------------------------

st.set_page_config(
        page_title="Planificador de Clases",
        page_icon="calendar",
    )

# Main app function
def main():
    st.title("Planificador de Clases 🎓")

    # Sidebar for id or University File Number
    st.sidebar.header("Información del usuario")
    id = st.sidebar.text_input("Ingresá tu número de legajo")

    # Save id to session state
    if id:
        st.session_state.id = id

    # Main app content
    st.markdown("""
        <style>
        html, body, [class*="st-"] {
            font-family: 'Gill Sans', 'Gill Sans MT', Calibri, 'Trebuchet MS', sans-serif;
        }
        .stButton>button {
            border: none;
            border-radius: px;
            padding: 10px 24px;
            margin: 5px 0px;
            color: white;
            background-color: #008CBA;
        }
        .stButton>button:hover {
            background-color: #005f73;
        }
        </style>
    """, unsafe_allow_html=True)

    st.markdown("Creado por [Nicolas Cantarovici](https://www.linkedin.com/in/nicolas-cantarovici-3b85a0198)")
    with st.expander("Como usar la app?"):
        st.markdown("""
            <div >
                <ol>
                    <li><strong>Registrá tus clases:</strong> Empezá ingresando todas las clases que te interesen. Ya sea las que asistas o que estés considerando, para ver todos los horarios posibles.</li>
                    <li><strong>Creá horarios posibles:</strong> Navegá a la pestaña 'Creador de horarios' para generar horarios personalizados basados en tus clases de interés. Experimentá con diferentes combinaciones para ajustar mejor los horarios.</li>
                    <li><strong>Descargá el archivo ICS:</strong> Después de finalizar de elegir las clases a cursar, ingresá a la sección 'Agregar al calendario'. Ahí, vas a poder descargar un archivo ICS, que se puede agregar fácilmente a tu calendario.</li>
                </ol>
            </div>
        """, unsafe_allow_html=True)



    # Create tabs for different functionalities
    tab1, tab2, tab3 = st.tabs(["📚 Registro de clases", "⏱️ Creador de horarios", "🗓️ Agregar al calendario"])

    with tab1:
        class_logger()

    with tab2:
        timetable_creator()

    with tab3:
        calendar_ics_generator()

def remove_class_from_db(id, class_name, group):
    collection.update_one(
        {"id": id},
        {"$pull": {"classes": {"name": class_name, "group": group}}}
    )

# Function to parse bulk schedule data
def parse_schedule_data(raw_data):
    # Split the raw data by lines
    lines = raw_data.strip().split('\n')
    
    # Initialize variables
    classes = {}
    current_class_name = None
    current_group = None

    for i in range(0, len(lines), 6):
        if len(lines[i:i+6]) < 6:
            continue
        
        class_name_match = re.match(r'^(.*)\s\(\d{4}\)', lines[i])
        if not class_name_match:
            continue
        class_name = class_name_match.group(1)
        group = lines[i+1].strip()
        day = lines[i+4].strip()
        time_range = lines[i+5].strip()
        
        if ' a ' not in time_range:
            continue
        
        start_time, end_time = time_range.split(' a ')
        
        try:
            start_time = datetime.strptime(start_time.strip(), "%H:%M").time()
            end_time = datetime.strptime(end_time.strip(), "%H:%M").time()
        except ValueError:
            continue
        
        if (class_name, group) not in classes:
            classes[(class_name, group)] = {
                "name": class_name,
                "group": group,
                "schedule": []
            }

        classes[(class_name, group)]['schedule'].append({
            "day": day,
            "start_time": start_time.strftime("%H:%M"),
            "end_time": end_time.strftime("%H:%M"),
            "class_room": "Sin asignar"  # Assuming classroom is "Sin asignar" based on provided data format
        })
    
    print( list(classes.values()))
    return list(classes.values())
        
# Streamlit app function
def class_logger():
    id = st.session_state.get('id')
    if not id:
        st.warning("Por favor, ingresá tu número de legajo en la barra lateral para registrar una clase.")
        return
    
    tab1, tab2, tab3 = st.tabs(["Añadir clase una por una", "Añadir clases en cantidad", "Eliminar clases"])

    with tab1:
        class_name = st.text_input("Nombre de la clase")
        group_section = st.text_input("Grupo/Sección")

        # Input for number of days and dynamic schedule inputs
        num_days = st.number_input("Número de días por semana", min_value=1, max_value=10, step=1, key='num_days')
        schedule_entries = []
        for i in range(num_days):
            cols = st.columns(4)
            with cols[0]:
                day = st.selectbox(f"Día {i+1}", ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes"], key=f"day{i}")
            with cols[1]:
                start_time = st.time_input(f"Hora de inicio {i+1}", key=f"start_time{i}", value=None)
            with cols[2]:
                end_time = st.time_input(f"Hora de fin {i+1}", key=f"end_time{i}", value=None)
            with cols[3]:
                class_room = st.text_input(f"Aula {i+1}", key=f"class_room{i}")
            schedule_entries.append((day, start_time, end_time, class_room))

        # Submission button
        submit_button = st.button("Registrar clase")

        # Handling the submission
        if submit_button and class_name and group_section:
            schedule = [
                {
                    "day": day, 
                    "start_time": start_time.strftime("%H:%M"), 
                    "end_time": end_time.strftime("%H:%M"),
                    "class_room": class_room
                } for day, start_time, end_time, class_room in schedule_entries
            ]
            
            new_class = {
                "name": class_name.capitalize(),
                "group": group_section,
                "schedule": schedule
            }

            save_class_to_db(id, new_class)

            display_class_entry(new_class)
            st.success("Clase registrada y guardada con éxito.")

    with tab2:
        st.write('Esta función es para los estudiantes de Ditella que pueden copiar y pegar las materias de [esta página](https://lookerstudio.google.com/u/1/reporting/56ee8002-2c18-425f-b3c9-5f107da7d0f8/page/hAzCD).')
        raw_data = st.text_area("Pega los datos de las clases aquí. (solo copiar y pegar el contenido de las materias)", height=300)
        submit_button_bulk = st.button("Registrar clases en bloque")

        if submit_button_bulk and raw_data:
            classes = parse_schedule_data(raw_data)
            for new_class in classes:
                save_class_to_db(id, new_class)
                # display_class_entry(new_class)
            st.success("Clases registradas y guardadas con éxito.")

    with tab3:
        classes = get_classes_from_db(id)
        if classes:
            class_names = [f"{cls['name']} - {cls['group']}" for cls in classes]
            selected_classes = st.multiselect("Selecciona las clases que deseas eliminar", class_names)
            if st.button("Eliminar clases"):
                for selected_class in selected_classes:
                    class_name, group = selected_class.split(' - ')
                    remove_class_from_db(id, class_name, group)
                st.success("Clases eliminadas con éxito")
        else:
            st.warning("No tienes clases registradas para eliminar")
            
        

# Timetable Creator tab
def timetable_creator():
    id = st.session_state.get('id')
    if id:
        classes = get_classes_from_db(id)
        if not classes:
            st.warning("No se encontraron clases para el ID ingresado. Por favor, registre algunas clases primero.")
            return

        parsed_classes = [{
            **cls,
            'schedule': [{
                **session,
                'start_time': parse_time(session['start_time']),
                'end_time': parse_time(session['end_time'])
            } for session in cls['schedule']]
        } for cls in classes]

        # Get unique class names
        class_names = list(set(cls['name'] for cls in parsed_classes))

        # Add multiselect for mandatory classes
        mandatory_classes = st.multiselect("Seleccioná las clases que quisieras probar en los posibles calendarios", class_names,placeholder='Elegir clases',max_selections=6)

        num_classes = st.number_input("Número de clases a asistir. No se pueden mas de 6 clases a la vez.", min_value=len(mandatory_classes), max_value=6, step=1)

        # User input for selecting free days
        days_of_week = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes']
        free_days = st.multiselect("Seleccioná los días en los que no querés tener clases", days_of_week,placeholder='Elegir días')

        generate_button = st.button("Generar opciones")

        if generate_button:
            try:
                class_combinations = combinations(parsed_classes, num_classes)
                viable_combinations = [
                    combo for combo in class_combinations
                    if not any(has_conflict(cls1, cls2) for cls1, cls2 in combinations(combo, 2))
                    and has_unique_classes(combo)
                    and has_free_days(combo, free_days)
                    and all(any(cls['name'] == mandatory_class for cls in combo) for mandatory_class in mandatory_classes)
                ]
                st.session_state['viable_combinations'] = viable_combinations

                if not viable_combinations:
                    st.warning("No se encontraron horarios con los criterios actuales. Considerá ajustar el número de clases, las clases obligatorias o los días libres seleccionados.")
                    return

                # Extract unique time slots and pass them along with the combinations and classes
                time_slots = get_unique_time_slots(parsed_classes)
                filename = 'horarios.xlsx'
                create_single_sheet_xlsx_timetables(viable_combinations, filename, time_slots, classes)


                # Provide download link
                st.success("Horario generado con éxito.")
                with open(filename, "rb") as file:
                    btn = st.download_button(
                            label="Descargar horario",
                            data=file,
                            file_name=filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
            except Exception as e:
                st.error(f"Ocurrió un error: {e}")
    else:
        st.warning("Por favor, ingresá tu número de legajo en la barra lateral para generar un horario.")

# Calendar ICS Generator tab
def calendar_ics_generator():
    id = st.session_state.get('id')
    if id:
        classes = get_classes_from_db(id)
        if not classes:
            st.warning("No se encontraron clases para el ID ingresado. Por favor, registrá algunas clases primero.")
            return

        # Combine class names with their sections for display
        class_display_names = [f"{cls['name']} - Sección {cls['group']}" for cls in classes]

        selected_classes_display = st.multiselect("Seleccioná las clases para incluir en el calendario definitivo", class_display_names,placeholder='Elegir clases')

        # Extract class names and groups from the selected display names
        selected_classes = [
            {'name': display.split(' - ')[0], 'group': display.split(' - ')[1].replace('Sección ', '')}
            for display in selected_classes_display
        ]

        if selected_classes:
            start_date = st.date_input("Fecha de inicio")
            end_date = st.date_input("Fecha de fin de cursada")

            if st.button("Generar calendario"):
                try:
                    ics_filename = generate_ics_file_for_classes(
                        selected_classes, classes, start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d')
                    )

                    with open(ics_filename, "rb") as file:
                        st.download_button(
                            label="Descargar archivo ICS",
                            data=file,
                            file_name=ics_filename,
                            mime="text/calendar"
                        )
                except Exception as e:
                    st.error(f"Ocurrió un error: {e}")
        else:
            st.warning("Por favor, seleccioná al menos una clase para incluir en el calendario.")
    else:
        st.warning("Por favor, ingresá tu número de legajo en la barra lateral para generar un calendario.")

if __name__ == "__main__":
    main()